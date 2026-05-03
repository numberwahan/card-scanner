from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)

_COL_WIDTHS = {
    "grader": 10, "grade": 10, "cert_number": 16, "year": 8,
    "player_name": 28, "set_name": 28, "card_number": 10,
    "variation": 20, "sport_category": 18, "purchase_price": 14,
    "value": 14, "sold_price": 14, "sold_date": 14,
    "date_scanned": 18, "image_path": 30,
}


def generate_excel(cards: list[dict], field_config: list[dict]) -> BytesIO:
    enabled = [f for f in field_config if f.get("enabled", True)]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Graded Cards"

    for col_idx, field in enumerate(enabled, start=1):
        cell = ws.cell(row=1, column=col_idx, value=field["label"])
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = _COL_WIDTHS.get(field["key"], 16)

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    for row_idx, card in enumerate(cards, start=2):
        for col_idx, field in enumerate(enabled, start=1):
            val = card.get(field["key"])
            if field["key"] == "date_scanned" and val:
                val = str(val)[:16]
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(vertical="center")
            if row_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="EBF3FB")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
